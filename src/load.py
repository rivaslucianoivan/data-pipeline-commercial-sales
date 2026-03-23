from pathlib import Path
import pandas as pd
from openpyxl.utils import get_column_letter

def asegurar_directorio_salida(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)


def exportar_a_excel(df_raw, df_clean, kpis: dict, tablas: dict, output_path) -> None:
    from pathlib import Path
    output_path = Path(output_path)
    asegurar_directorio_salida(output_path.parent)

    with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        # 1) Hoja raw
        sheet_name = "raw"
        df_raw.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        _formatear_hoja_basica(ws, df_raw)

        # 2) Hoja clean
        sheet_name = "clean"
        df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        _formatear_hoja_basica(ws, df_clean)

        # 3) Hojas de KPIs
        for nombre, df_kpi in kpis.items():
            sheet_name = f"kpi_{nombre}"
            df_kpi.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _formatear_hoja_basica(ws, df_kpi)

        # 4) Hojas de tablas resumen
        for nombre, df_tab in tablas.items():
            sheet_name = f"tbl_{nombre}"
            df_tab.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _formatear_hoja_basica(ws, df_tab)

def _formatear_hoja_basica(worksheet, df):
    # Congelar encabezados
    worksheet.freeze_panes = "A2"

    # Auto-filtro en toda la tabla (desde A1 hasta última fila/columna)
    if len(df.columns) > 0 and len(df) > 0:
        last_col_letter = get_column_letter(len(df.columns))
        last_row = len(df) + 1  # +1 por el encabezado
        ref = f"A1:{last_col_letter}{last_row}"
        worksheet.auto_filter.ref = ref

    # Ajuste de ancho de columnas...
    for idx, col in enumerate(df.columns):
        serie = df[col].astype(str)
        max_len = max(serie.map(len).max(), len(str(col))) + 2
        if max_len > 50:
            max_len = 50
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

